# -*- coding: utf-8 -*-
#####################################################
#    Project:     2 Line Caller ID & Cus. Details   #
#    Programmer:  Sina Shiri                        #
#    Date:        2017 Sep 21                       #
#    For:         Getir Market Co.                  #
#####################################################

# import needed Packages
from tkinter import font
from tkinter import *
import jdatetime
import serial
from os.path import exists
from os import system
import os
import subprocess
from distutils.dir_util import copy_tree
from openpyxl import Workbook
from openpyxl import load_workbook

# Public Variables

#########################################
# Create and Configure GUI with Tkinter #
#########################################
root = Tk()
root.title("RollCall with RFID on RPi")
### Get Screen Resolution and Create Full screen GUI
screen_x = root.winfo_screenwidth()  # for Getirmarket PC : 1024
screen_y = root.winfo_screenheight()  # for Getirmarket PC : 0768
# root.overrideredirect(True)
root.resizable(0, 0)
root.geometry('%dx%d+%d+%d' % (screen_x / 4, screen_y / 4, screen_x / 1.37, screen_y / 1.536))
### Create Canvas for Design GUI
w = Canvas(root, width=screen_x / 4, height=screen_y / 4, background='violet')
w.grid(row=0, column=0)
### Put a High Resolution Pic in GUI as Background
background_image = PhotoImage(file="images\\back_g.png")
w.create_image(0, 0, anchor=NW, image=background_image)

### Create Fonts and use in GUI
### Persian Fonts
status_font = font.Font(family="B Titr", size=15, weight="bold")
head_font = font.Font(family="B Zar", size=20, weight="bold")
field_font = font.Font(family="B Zar", size=15)
infiled_font = font.Font(family="B Zar", size=12, weight="bold")
detail_font = font.Font(family="B Zar", size=14, weight="bold")
list_font = font.Font(family="B Zar", size=12, weight="bold")



##################
# MAIN FUNCTIONs #
##################
def show_customer(line_num, customer_row):
    show_c_win = Toplevel(root)
    show_c_win.title("Known Customer")
    show_c_win.resizable(0, 0)
    show_c_win.geometry('%dx%d+%d+%d' % (screen_x / 2.6, screen_y / 4, screen_x / 3, screen_y / 1.536))
    show_c_canvas = Canvas(show_c_win, width=screen_x / 2, height=screen_y / 2.5, background='violet')
    show_c_canvas.grid(row=0, column=0)
    global show_c_backgnd
    show_c_backgnd = PhotoImage(file="images\\back_add.png")
    show_c_canvas.create_image(0, 0, anchor=NW, image=show_c_backgnd)
    ### Constant Texts
    show_c_canvas.create_text(screen_x / 3, screen_y / 30, text=":مشتری", font=head_font, fill='black')
    show_c_canvas.create_text(screen_x / 3, screen_y / 10, text=":کد مـشترک", font=head_font, fill='black')
    show_c_canvas.create_text(screen_x / 3, screen_y / 6.8, text=":آدرــــــس", font=head_font, fill='black')
    show_c_canvas.create_text(screen_x / 3, screen_y / 5.1, text=":شماره همراه", font=head_font, fill='black')
    if line_num == 1:
        show_c_canvas.create_text(screen_x / 13, screen_y / 30, text="(1622)خط: خط یک", font=head_font, fill='black')
    else:
        show_c_canvas.create_text(screen_x / 13, screen_y / 30, text="(1664)خط: خط دو", font=head_font, fill='black')

    ### Variable Texts
    ##### Get Cus. Details from Xlsx file
    wb = load_workbook("Data\\Sample.xlsx")
    ws = wb.active
    ### Show Texts
    show_c_canvas.create_text(screen_x / 4, screen_y / 30, text=ws["B"+str(customer_row)].value,
                              font=detail_font, justify='right', fill='black')
    show_c_canvas.create_text(screen_x / 3.7, screen_y / 10, text=ws["A" + str(customer_row)].value,
                              font=detail_font, justify='right', fill='black')
    show_c_canvas.create_text(screen_x / 7, screen_y / 6.8, text=ws["D" + str(customer_row)].value,
                              font=detail_font, justify='right', fill='black')
    show_c_canvas.create_text(screen_x / 4, screen_y / 5.1, text=ws["C" + str(customer_row)].value,
                              font=detail_font, justify='right', fill='black')
def check_number(number):
    wb = load_workbook("Data\\Sample.xlsx")
    ws = wb.active
    customer_cnt_file = open("Data\\customer_cnt.txt")
    customer_cnt = int(customer_cnt_file.read()[18:])
    customer_cnt_file.close()
    for i in range(1, customer_cnt - 150, 1):
        num_ = ws["C" + str(i)]
        if str(num_.value) == str(number):
            return "Find", i
    return "Not Find", 0


def add_person():
    global customer_cnt
    ### Open Xlsx file and save new Customer
    wb = load_workbook("Data\\Data.xlsx")
    ws = wb.active
    ws["A" + str(customer_cnt - 150)] = customer_cnt
    ws["B" + str(customer_cnt-150)] = user_name.get()
    ws["C" + str(customer_cnt - 150)] = user_phone.get()
    ws["D" + str(customer_cnt - 150)] = user_address.get()
    wb.save("Data\\Data.xlsx")

    ### Increase customer_cnt Value
    customer_cnt = customer_cnt + 1
    customer_cnt_file = open("Data\\customer_cnt.txt", 'w')
    customer_cnt_file.write("total_of_customer=" + str(customer_cnt))
    customer_cnt_file.close()

def add_fun():
    add_win = Toplevel(root)
    add_win.title("Add New Customer")
    add_win.resizable(0, 0)
    add_win.geometry('%dx%d+%d+%d' % (screen_x / 2, screen_y / 2.5, screen_x / 5, screen_y / 7.68))
    add_canvas = Canvas(add_win, width=screen_x / 2, height=screen_y / 2.5, background='violet')
    add_canvas.grid(row=0, column=0)
    global add_backgnd
    add_backgnd = PhotoImage(file="images\\back_add.png")
    add_canvas.create_image(0, 0, anchor=NW, image=add_backgnd)

    ##### Generate New Code and Show in GUI
    global customer_cnt
    customer_cnt_file = open("Data\\customer_cnt.txt")
    customer_cnt = int(customer_cnt_file.read()[18:])
    customer_cnt_file.close()
    add_canvas.create_text(screen_x / 2.8, screen_y / 30 + screen_y / 15,
                           text=str(customer_cnt), font=field_font, fill='white')
    ### Texts
    add_canvas.create_text(screen_x/2.4, screen_y/30, text="افزودن مشتری", font=head_font, fill='black')
    add_canvas.create_text(screen_x / 2.3, screen_y /30 + screen_y/15,
                           text=":کد اشتراک", font=field_font, fill='white')
    add_canvas.create_text(screen_x / 2.3, screen_y /30 + 2*screen_y/15,
                           text="نام مشترک", font=field_font, fill='white')
    add_canvas.create_text(screen_x / 2.3, screen_y /30 + 3*screen_y/15 ,
                           text="آدرس", font=field_font, fill='white')
    add_canvas.create_text(screen_x / 2.3, screen_y / 30 + 4 * screen_y / 15,
                           text="شماره همراه", font=field_font, fill='white')
    ### Button
    add_button = Button(add_win, command=add_person)
    add_button.configure(image=add_person_icon, width=screen_x/20, height=screen_y/20, background="red", relief=FLAT)
    add_canvas.create_window(screen_x/30,6.5*screen_y/20, anchor=NW, window=add_button)
    ### Fields
    global user_name, user_address, user_phone
    user_name = Entry(add_canvas, font=infiled_font, justify="right", state=NORMAL)
    add_canvas.create_window(screen_x / 3.5,screen_y /30 + 2*screen_y/15,
                             width=screen_x / 5, height=screen_y / 23, window=user_name)
    user_address = Entry(add_canvas, font=infiled_font, justify="right", state=NORMAL)
    add_canvas.create_window(screen_x / 5, screen_y / 30 + 3 * screen_y / 15,
                             width= 1.87*screen_x / 5, height=screen_y / 23, window=user_address)
    user_phone = Entry(add_canvas, font=infiled_font, justify="right", state=NORMAL)
    add_canvas.create_window(screen_x / 3.5, screen_y / 30 + 4 * screen_y / 15,
                             width=screen_x / 5, height=screen_y / 23, window=user_phone)

def view_fun():
    view_win = Toplevel(root)
    view_win.title("View Customers List")
    view_win.resizable(0, 0)
    view_win.geometry('%dx%d+%d+%d' % (screen_x / 2, screen_y / 2.5, screen_x / 5, screen_y / 7.68))
    view_canvas = Canvas(view_win, width=screen_x / 2, height=screen_y / 2.5, background='violet')
    view_canvas.grid(row=0, column=0)
    global view_backgnd
    view_backgnd = PhotoImage(file="images\\back_add.png")
    view_canvas.create_image(0, 0, anchor=NW, image=view_backgnd)

    ### Texts
    view_canvas.create_text(screen_x / 2.4, screen_y / 30, text="مشتریان", font=head_font, fill='black')
    view_canvas.create_text(screen_x / 2.3, screen_y / 30 + screen_y / 20,
                           text="جستجو", font=field_font, fill='white')
    ### Buttons
    del_button_view = Button(view_win, command=del_fun)
    del_button_view.configure(image=del_person_icon, width=35, height=35, background="red", relief=FLAT)
    view_canvas.create_window(screen_x/90, screen_y/90, anchor=NW, window=del_button_view)

    ### Field
    global search_
    search_ = Entry(view_canvas, font=infiled_font, justify="right", state=NORMAL)
    view_canvas.create_window(screen_x / 3.4, screen_y / 30 + screen_y / 19,
                             width=screen_x / 5, height=screen_y / 26, window=search_)
    ### add List of Customers
    global remove_list
    remove_list = Listbox(view_win, font=list_font)
    global scroll
    scroll = Scrollbar(view_win, orient=VERTICAL, command=remove_list.yview)
    remove_list['yscrollcommand'] = scroll.set
    global remove_list_box
    remove_list_box = view_canvas.create_window(screen_x/4, screen_y/4,
                                      width=screen_x /3, height=screen_y / 4, window=remove_list)
    global scroll_box
    scroll_box = view_canvas.create_window(screen_x/2.35 ,screen_y/4,
                                 width=screen_x /50, height=screen_y / 4, window=scroll)
    ### insert List members
    remove_list.insert('end',  )

def del_fun():
    print("del")


#####################
# Set main sections #
#####################
def set_main_win():
    ### texts
    call_status = w.create_text(184, 80, text=": تـماس امـروز", font=status_font, fill='black')
    L1_status = w.create_text(185, 120, text=":وضعیت خط یک", font=status_font, fill='black')
    L2_status = w.create_text(185, 160, text=": وضعیت خط دو", font=status_font, fill='black')

    ### Buttons
    global add_person_icon
    add_person_icon = PhotoImage(file="images\\add_person.png")
    add_button = Button(root, command=add_fun)
    add_button.configure(image=add_person_icon, width=35, height=35, background="white", relief=FLAT)
    add_button_window = w.create_window(16, 7, anchor=NW, window=add_button)

    global view_person_icon
    view_person_icon = PhotoImage(file="images\\view.png")
    view_button = Button(root, command=view_fun)
    view_button.configure(image=view_person_icon, width=35, height=35, background="white", relief=FLAT)
    view_button_window = w.create_window(70, 9, anchor=NW, window=view_button)

    global del_person_icon
    del_person_icon = PhotoImage(file="images\\del_person.png")
    del_button = Button(root, command=del_fun)
    del_button.configure(image=del_person_icon, width=35, height=35, background="white", relief=FLAT)
    del_button_window = w.create_window(120, 7, anchor=NW, window=del_button)





def main__loop():
    line_status = input("INPUT:")
    if line_status[0:4] == "ring":
        number_known, customer_row = check_number(line_status[4:])
        if number_known == "Find":
            line_num = 2
            show_customer(line_num, customer_row)
        elif number_known == "Not Find":
            add_ring_num()



set_main_win()
root.after(100, main__loop)
root.mainloop()
